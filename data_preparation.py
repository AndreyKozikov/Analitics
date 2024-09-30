import pandas as pd
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor


# Функции для обработки данных
def calculate_formula(row):
    currency_symbol = row['Валюта']  # Символ валюты (например, $)
    if currency_symbol == '$':  # Если символ "$", используем курс USD
        currency_code = 'USD'
    else:
        currency_code = 'RUB'  # Для других валют (например, рубли) ставим курс 1

    if currency_code in currency_dict:
        price = float(row['Цена'])
        rate = float(currency_dict[currency_code].replace(",", "."))
        return price * rate
    return row['Цена']  # Если курс не найден, вернуть цену как есть


def calculate_margin(row):
    return row['Маржинальность'] * row['Итоговая стоимость в рублях']


def touch_chains(marketing_data):
    marketing_data['Device Category'] = marketing_data['Device Category'].astype(str)
    touch_chains = marketing_data.groupby('Client ID')['Device Category'].apply(lambda x: ' -> '.join(x)).reset_index()
    conversion_sum = marketing_data.groupby('Client ID')['Конверсия'].sum().reset_index()
    conversion_sum['Признак конверсии'] = conversion_sum['Конверсия'].apply(lambda x: 1 if x > 0 else 0)
    touch_count = marketing_data.groupby('Client ID').size().reset_index(name='Число касаний')
    final_data = touch_chains.merge(conversion_sum, on='Client ID').merge(touch_count, on='Client ID')
    final_data.rename(columns={'Конверсия': 'Сумма конверсий по цепочкам'}, inplace=True)
    return "Цепочки касаний", final_data


def calculate_price(reference_data):
    reference_data['Итоговая стоимость в рублях'] = reference_data.apply(calculate_formula, axis=1)
    reference_data['Маржинальность в рублях'] = reference_data.apply(calculate_margin, axis=1)
    return "Справочник", reference_data


def model_add(model_dict1, marketing_data_sheet):
    # Приведение столбцов к строковому типу
    marketing_data_sheet['Domain'] = marketing_data_sheet['Domain'].astype(str)
    marketing_data_sheet['Goal Completion Location'] = marketing_data_sheet['Goal Completion Location'].astype(str)

    # Создаем маски для обновлений
    marka_mask = marketing_data_sheet['Марка'].copy()
    model_mask = marketing_data_sheet['Модель'].copy()

    for model, marka in model_dict1.items():

        # Проверяем, что marka и model - это строки и не пустые
        if isinstance(model, str) and model:
            mask = (marketing_data_sheet['Domain'].str.contains(model, case=False, na=False) |
                    marketing_data_sheet['Goal Completion Location'].str.contains(model, case=False, na=False))
            model_mask[mask] = model

        if isinstance(marka, str) and marka:
            mask = (marketing_data_sheet['Domain'].str.contains(marka, case=False, na=False) |
                    marketing_data_sheet['Goal Completion Location'].str.contains(marka, case=False, na=False))
            marka_mask[mask] = marka

    # Обновляем столбцы в датафрейме
    marketing_data_sheet['Модель'] = model_mask
    marketing_data_sheet['Марка'] = marka_mask

    return "Маркетинговые данные", marketing_data_sheet


def clean_marka(marka):
    # Убираем все пробелы и приводим к нижнему регистру
    cleaned_marka = marka.strip().replace(" ", "").lower()
    # Если марка BMW, оставляем все буквы заглавными
    if cleaned_marka == 'bmw' or cleaned_marka == 'bмw':
        return 'BMW'
    if cleaned_marka == 'bmw' or cleaned_marka == 'bмw':
        return 'BMW'
    # Приводим первую букву к заглавной, остальные к строчным
    return cleaned_marka.capitalize()

file_path = 'processed_data.xlsx'

# Загружаем существующий файл Excel
workbook = load_workbook(file_path)

# Извлечение нужных листов
reference = pd.DataFrame(workbook['Справочник'].values)
reference.columns = reference.iloc[0]  # Установка заголовков
reference = reference[1:]  # Удаление первой строки с заголовками
reference['Марка'] = reference['Марка'].astype(str).apply(clean_marka)

marketing_data = pd.DataFrame(workbook['Маркетинговые данные'].values)
marketing_data.columns = marketing_data.iloc[0]
marketing_data = marketing_data[1:]
# Заменяем 'Mersedes' на 'Mercedes' в столбце 'Domain'
marketing_data['Domain'] = marketing_data['Domain'].str.replace('Mersedes', 'Mercedes', case=False)
marketing_data['Марка'] = ''
marketing_data['Модель'] = ''

currency_data = pd.DataFrame(workbook['Курсы валют'].values)
currency_data.columns = currency_data.iloc[0]
currency_data = currency_data[1:]

# Создание словаря курсов валют
currency_dict = dict(zip(currency_data['Букв. код'], currency_data['Курс']))

# Создание словаря курсов валют
model_dict = dict(zip(reference['Модель'], reference['Марка']))
print(model_dict)

# Использование многопоточности для выполнения функций параллельно
with ThreadPoolExecutor() as executor:
    futures = {
        'marketing_data': executor.submit(model_add, model_dict, marketing_data),
        'reference': executor.submit(calculate_price, reference)
    }

    # Получаем результаты из futures
    marketing_data_name, marketing_data = futures['marketing_data'].result()
    reference_name, reference = futures['reference'].result()

# Теперь, когда marketing_data уже преобразован, мы можем вызвать touch_chains
touch_chains_name, touch_chains = touch_chains(marketing_data)

# Удаляем старые листы, если они существуют
for sheet_name in [reference_name, marketing_data_name, touch_chains_name]:
    if sheet_name in workbook.sheetnames:
        std = workbook[sheet_name]
        workbook.remove(std)
workbook.save(file_path)


# Добавляем новые данные в файл Excel
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
    # Записываем новые листы
    reference.to_excel(writer, sheet_name=reference_name, index=False)
    marketing_data.to_excel(writer, sheet_name=marketing_data_name, index=False)
    touch_chains.to_excel(writer, sheet_name=touch_chains_name, index=False)


